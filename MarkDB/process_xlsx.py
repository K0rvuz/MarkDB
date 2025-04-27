from pathlib import Path
import datetime
import sqlite3
from db import DB_PATH
import openpyxl
from utils import format_table
import streamlit as st




# === PROCESSADOR DE XLSX ===
def process_xlsx_to_chunks(file_path):
    """Processa arquivos XLSX (Excel)"""
    file_name = Path(file_path).name
    upload_time = datetime.now().strftime("%d/%m/%Y")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        sheet_counter = 0
        
        for sheet_name in wb.sheetnames:
            sheet_counter += 1
            ws = wb[sheet_name]
            data = []
            
            for row in ws.iter_rows(values_only=True):
                data.append(row)
            
            if data:
                markdown_table = format_table(data)
                if markdown_table:
                    markdown_chunk = f"**Arquivo**: {file_name}\n**Planilha**: {sheet_name}\n**Chunk**: {sheet_counter}\n\n{markdown_table}"
                    c.execute('''
                        INSERT INTO chunks (file_name, page_number, chunk_number, content, chunk_content, upload_time, image_description, image_data)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (file_name, sheet_counter, 1, markdown_chunk, "tabela", upload_time, None, None))
        
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        st.error(f"Erro ao processar planilha Excel: {str(e)}")
        return False
    

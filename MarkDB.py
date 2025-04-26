import sqlite3
import os
import base64
from pathlib import Path
import fitz  # PyMuPDF
import textwrap
import docx
import openpyxl
import streamlit as st
import pandas as pd
from io import StringIO, BytesIO
from datetime import datetime
from openai import OpenAI
from PIL import Image

# === CONFIGURAÇÕES ===
if 'CHUNK_SIZE' not in st.session_state:
    st.session_state.CHUNK_SIZE = 800  # Tamanho padrão dos chunks de texto
DB_PATH = "markdb.sqlite"  # Caminho do banco de dados

# === INICIALIZAÇÃO DO BANCO DE DADOS ===
def init_db():
    """Inicializa o banco de dados SQLite com a estrutura necessária"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS chunks (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_name TEXT,
            page_number INTEGER,
            chunk_number INTEGER,
            content TEXT,
            chunk_content TEXT,
            upload_time TEXT,
            image_description TEXT,
            image_data TEXT
        )
    ''')
    conn.commit()
    conn.close()

# === FUNÇÕES AUXILIARES ===
def get_openai_client():
    """Retorna o cliente OpenAI configurado"""
    # Primeiro verifica se há uma chave no session_state (inserida pelo usuário)
    if hasattr(st.session_state, 'openai_api_key') and st.session_state.openai_api_key:
        return OpenAI(api_key=st.session_state.openai_api_key)
    # Depois verifica secrets.toml (para deploy)
    try:
        if 'openai_api_key' in st.secrets:
            return OpenAI(api_key=st.secrets.openai_api_key)
    except:
        pass
    return None

def generate_image_description(image_bytes):
    """Gera descrição de imagem usando a API da OpenAI"""
    client = get_openai_client()
    if client is None:
        st.error("API key da OpenAI não configurada")
        return None
    
    try:
        b64_image = base64.b64encode(image_bytes).decode('utf-8')
        response = client.chat.completions.create(
            model="gpt-4-turbo",
            messages=[
                {
                    "role": "user",
                    "content": [
                        {"type": "text", "text": "Descreva esta imagem detalhadamente em português."},
                        {
                            "type": "image_url",
                            "image_url": {
                                "url": f"data:image/png;base64,{b64_image}"
                            },
                        },
                    ],
                }
            ],
            max_tokens=500,
        )
        return response.choices[0].message.content
    except Exception as e:
        st.error(f"Erro ao gerar descrição: {str(e)}")
        return None

def detect_chunk_type(chunk):
    """Detecta o tipo de conteúdo do chunk"""
    if "```" in chunk:  # Se contém blocos de código
        return "código"
    elif any(char.isdigit() for char in chunk) and sum(c.isalpha() for c in chunk) / len(chunk) < 0.5:
        return "dados numéricos"
    elif len(chunk.split('\n')) > 5 and all(len(line) < 100 for line in chunk.split('\n')):
        return "lista"
    else:
        return "texto"

def format_table(matrix):
    """Formata uma matriz como tabela Markdown"""
    if not matrix or len(matrix) < 2:
        return ""
    
    # Cabeçalho
    markdown = "| " + " | ".join(str(cell) for cell in matrix[0]) + " |\n"
    markdown += "| " + " | ".join(["---"] * len(matrix[0])) + " |\n"
    
    # Linhas
    for row in matrix[1:]:
        markdown += "| " + " | ".join(str(cell) for cell in row) + " |\n"
    
    return markdown

# === PROCESSADOR DE IMAGENS ===
def process_image_to_chunks(file_path):
    """Processa arquivos de imagem (JPG, PNG)"""
    file_name = Path(file_path).name
    
    # Verifica se temos um cliente OpenAI configurado
    client = get_openai_client()
    if client is None:
        st.error("""🔒 API da OpenAI não configurada. 
                Por favor, vá para 'Configurações' e insira sua chave API para gerar descrições de imagens.""")
        return False
    
    try:
        with open(file_path, "rb") as f:
            image_bytes = f.read()
        
        # Gera descrição da imagem
        with st.spinner("Gerando descrição da imagem..."):
            image_description = generate_image_description(image_bytes)
        
        if image_description is None:
            st.warning("Não foi possível gerar uma descrição para a imagem")
            image_description = "Descrição não disponível"
        
        # Codifica a imagem em base64
        b64_image = base64.b64encode(image_bytes).decode('ascii')
        
        # Prepara o conteúdo para o banco de dados
        image_meta = f"**Arquivo**: {file_name}\n**Tipo**: Imagem\n**Descrição**: {image_description}"
        
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('''
            INSERT INTO chunks (file_name, page_number, chunk_number, content, chunk_content, upload_time, image_description, image_data)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (file_name, 1, 1, image_meta, "imagem", datetime.now().strftime("%d/%m/%Y"), image_description, b64_image))
        
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        st.error(f"Erro ao processar imagem: {str(e)}")
        return False

# === PROCESSADOR DE PDF ===
def process_pdf_to_chunks(file_path):
    """Processa arquivos PDF e extrai texto, tabelas e imagens"""
    doc = fitz.open(file_path)
    file_name = Path(file_path).name
    upload_time = datetime.now().strftime("%d/%m/%Y")

    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    for page_index in range(len(doc)):
        page = doc.load_page(page_index)
        text = page.get_text("text")
        tables = page.find_tables()
        images = page.get_images(full=True)
        chunk_counter = 1

        # Processa texto
        if text.strip():
            chunks = textwrap.wrap(text, st.session_state.CHUNK_SIZE)
            for chunk in chunks:
                chunk_type = detect_chunk_type(chunk)
                markdown_chunk = f"**Arquivo**: {file_name}\n**Página**: {page_index + 1}\n**Chunk**: {chunk_counter}\n\n{chunk.strip()}"
                c.execute('''
                    INSERT INTO chunks (file_name, page_number, chunk_number, content, chunk_content, upload_time, image_description, image_data)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (file_name, page_index + 1, chunk_counter, markdown_chunk, chunk_type, upload_time, None, None))
                chunk_counter += 1

        # Processa tabelas
        if tables:
            for table in tables:
                matrix = table.extract()
                markdown_table = format_table(matrix)
                if markdown_table:
                    markdown_chunk = f"**Arquivo**: {file_name}\n**Página**: {page_index + 1}\n**Chunk**: {chunk_counter}\n\n{markdown_table}"
                    c.execute('''
                        INSERT INTO chunks (file_name, page_number, chunk_number, content, chunk_content, upload_time, image_description, image_data)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (file_name, page_index + 1, chunk_counter, markdown_chunk, "tabela", upload_time, None, None))
                    chunk_counter += 1

        # Processa imagens
        for img in images:
            xref = img[0]
            base_image = doc.extract_image(xref)
            image_bytes = base_image["image"]
            
            try:
                # Verifica se temos cliente OpenAI para gerar descrição
                client = get_openai_client()
                if client is None:
                    image_description = "Descrição não disponível (API não configurada)"
                else:
                    # Gera descrição da imagem
                    image_description = generate_image_description(image_bytes)
                    if image_description is None:
                        image_description = "Descrição não disponível"
                
                # Codifica a imagem em base64
                b64_image = base64.b64encode(image_bytes).decode('ascii')
                
                # Armazena metadados e dados da imagem
                image_meta = f"**Arquivo**: {file_name}\n**Página**: {page_index + 1}\n**Chunk**: {chunk_counter}\n**Descrição**: {image_description}"
                
                c.execute('''
                    INSERT INTO chunks (file_name, page_number, chunk_number, content, chunk_content, upload_time, image_description, image_data)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (file_name, page_index + 1, chunk_counter, image_meta, "imagem", upload_time, image_description, b64_image))
                chunk_counter += 1
                
            except Exception as e:
                st.error(f"Erro ao processar imagem: {str(e)}")
                continue

    conn.commit()
    conn.close()

# === PROCESSADOR DE DOCX ===
def process_docx_to_chunks(file_path):
    """Processa arquivos DOCX (Word)"""
    file_name = Path(file_path).name
    upload_time = datetime.now().strftime("%d/%m/%Y")
    
    try:
        doc = docx.Document(file_path)
        full_text = []
        
        for para in doc.paragraphs:
            full_text.append(para.text)
        
        text = '\n'.join(full_text)
        
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        
        if text.strip():
            chunks = textwrap.wrap(text, st.session_state.CHUNK_SIZE)  # Usando session_state
            for i, chunk in enumerate(chunks, 1):
                chunk_type = detect_chunk_type(chunk)
                markdown_chunk = f"**Arquivo**: {file_name}\n**Chunk**: {i}\n\n{chunk.strip()}"
                c.execute('''
                    INSERT INTO chunks (file_name, page_number, chunk_number, content, chunk_content, upload_time, image_description, image_data)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ''', (file_name, 1, i, markdown_chunk, chunk_type, upload_time, None, None))
        
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        st.error(f"Erro ao processar documento Word: {str(e)}")
        return False

# === PROCESSADOR DE XLSX ===
def process_xlsx_to_chunks(file_path):
    """Processa arquivos XLSX (Excel)"""
    file_name = Path(file_path).name
    upload_time = datetime.now().strftime("%d/%m/%Y")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        sheet_counter = 0
        
        for sheet_name in wb.sheetnames:
            sheet_counter += 1
            ws = wb[sheet_name]
            data = []
            
            for row in ws.iter_rows(values_only=True):
                data.append(row)
            
            if data:
                markdown_table = format_table(data)
                if markdown_table:
                    markdown_chunk = f"**Arquivo**: {file_name}\n**Planilha**: {sheet_name}\n**Chunk**: {sheet_counter}\n\n{markdown_table}"
                    c.execute('''
                        INSERT INTO chunks (file_name, page_number, chunk_number, content, chunk_content, upload_time, image_description, image_data)
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    ''', (file_name, sheet_counter, 1, markdown_chunk, "tabela", upload_time, None, None))
        
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        st.error(f"Erro ao processar planilha Excel: {str(e)}")
        return False

# === EXPORTAÇÃO PARA MD ===
def export_chunk_to_md(chunk_id):
    """Exporta um chunk específico para um arquivo Markdown"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()
    c.execute("SELECT * FROM chunks WHERE id = ?", (chunk_id,))
    row = c.fetchone()
    conn.close()
    
    if not row:
        return None, None
    
    columns = ['id', 'file_name', 'page_number', 'chunk_number', 'content', 
              'chunk_content', 'upload_time', 'image_description', 'image_data']
    chunk_data = dict(zip(columns, row))
    
    # Constrói o conteúdo Markdown
    md_content = f"# {chunk_data['file_name']}\n\n"
    md_content += f"**ID:** {chunk_data['id']}\n"
    md_content += f"**Página:** {chunk_data['page_number']}\n"
    md_content += f"**Chunk:** {chunk_data['chunk_number']}\n"
    md_content += f"**Tipo:** {chunk_data['chunk_content']}\n"
    md_content += f"**Data de upload:** {chunk_data['upload_time']}\n\n"
    
    if chunk_data['chunk_content'] == "imagem":
        md_content += "## Descrição da Imagem\n"
        md_content += f"{chunk_data['image_description']}\n\n"
        md_content += f"![Imagem](data:image/png;base64,{chunk_data['image_data']})"
    else:
        md_content += chunk_data['content']
    
    return md_content, f"chunk_{chunk_id}_{chunk_data['file_name']}.md"

# === INTERFACE STREAMLIT ===
def main():
    st.set_page_config(page_title="MarkDB", page_icon="📄", layout="wide")
    st.title("📄 MarkDB - Banco de Dados Markdown")
    
    init_db()
    
    # Configuração inicial da API Key
    if 'openai_api_key' not in st.session_state:
        st.session_state.openai_api_key = ""
    
    menu = st.sidebar.selectbox("Navegar", ["Upload de Arquivo", "Banco de Dados", "Configurações"])

    if menu == "Upload de Arquivo":
        st.header("📤 Upload de Arquivo")
        uploaded_file = st.file_uploader("Escolha um arquivo (.pdf, .docx, .xlsx, .jpg, .png)", 
                                       type=["pdf", "docx", "xlsx", "jpg", "jpeg", "png"])
        
        if uploaded_file is not None:
            file_name = uploaded_file.name
            file_ext = file_name.split('.')[-1].lower()
            temp_dir = "temp_upload"
            os.makedirs(temp_dir, exist_ok=True)
            file_path = os.path.join(temp_dir, file_name)
            
            with open(file_path, "wb") as f:
                f.write(uploaded_file.getbuffer())
            
            if file_ext in ["jpg", "jpeg", "png"]:
                success = process_image_to_chunks(file_path)
                if success:
                    st.success("✅ Imagem processada com sucesso!")
                else:
                    st.error("❌ Falha ao processar a imagem")
            elif file_ext == "pdf":
                with st.spinner("Processando PDF..."):
                    process_pdf_to_chunks(file_path)
                    st.success("✅ PDF processado com sucesso!")
            elif file_ext == "docx":
                with st.spinner("Processando documento Word..."):
                    process_docx_to_chunks(file_path)
                    st.success("✅ Documento Word processado com sucesso!")
            elif file_ext == "xlsx":
                with st.spinner("Processando planilha Excel..."):
                    process_xlsx_to_chunks(file_path)
                    st.success("✅ Planilha Excel processada com sucesso!")
            
            if os.path.exists(file_path):
                os.remove(file_path)

    elif menu == "Banco de Dados":
        st.header("🔎 Banco de Dados")
        conn = sqlite3.connect(DB_PATH)
        df = pd.read_sql_query("SELECT * FROM chunks", conn)
        conn.close()

        if df.empty:
            st.info("ℹ️ O banco de dados está vazio. Faça o upload de um arquivo para começar.")
        else:
            st.write("### 📊 Conteúdo do Banco de Dados")
            
            # Filtros
            col1, col2, col3 = st.columns(3)
            with col1:
                filter_file = st.selectbox("Filtrar por arquivo:", ["Todos"] + list(df['file_name'].unique()))
            with col2:
                filter_type = st.selectbox("Filtrar por tipo:", ["Todos"] + list(df['chunk_content'].unique()))
            with col3:
                items_per_page = st.selectbox("Itens por página:", [10, 25, 50, 100])

            # Aplicar filtros
            filtered_df = df.copy()
            if filter_file != "Todos":
                filtered_df = filtered_df[filtered_df['file_name'] == filter_file]
            if filter_type != "Todos":
                filtered_df = filtered_df[filtered_df['chunk_content'] == filter_type]

            # Paginação
            total_pages = max(1, (len(filtered_df) // items_per_page) + 1)
            page_number = st.number_input("Página:", min_value=1, max_value=total_pages, value=1)
            start_idx = (page_number - 1) * items_per_page
            end_idx = start_idx + items_per_page

            # Exibir dados filtrados
            st.dataframe(filtered_df.iloc[start_idx:end_idx], height=400, use_container_width=True)

            # Visualização detalhada
            st.write("### 🔍 Visualização Detalhada")
            chunk_id = st.number_input("ID do chunk para visualizar:", min_value=1, max_value=df['id'].max(), value=1)
            
            if st.button("Visualizar Chunk"):
                conn = sqlite3.connect(DB_PATH)
                c = conn.cursor()
                c.execute("SELECT * FROM chunks WHERE id = ?", (chunk_id,))
                row = c.fetchone()
                conn.close()
                
                if row:
                    columns = ['id', 'file_name', 'page_number', 'chunk_number', 'content', 
                              'chunk_content', 'upload_time', 'image_description', 'image_data']
                    chunk_data = dict(zip(columns, row))
                    
                  
                    
                    st.write("#### Conteúdo")
                    if chunk_data['chunk_content'] == "imagem":
                        try:
                            image_bytes = base64.b64decode(chunk_data['image_data'])
                            st.image(BytesIO(image_bytes), caption=chunk_data['file_name'], use_container_width=True)
                            st.write(chunk_data['content'])
                            
                        except Exception as e:
                            st.error(f"Erro ao exibir imagem: {str(e)}")
                    else:
                        st.markdown(chunk_data['content'])
                    
                    # Exportação para MD
                    md_content, md_filename = export_chunk_to_md(chunk_id)
                    st.download_button(
                        label="⬇️ Baixar como Markdown",
                        data=md_content,
                        file_name=md_filename,
                        mime="text/markdown"
                    )
                else:
                    st.warning("Chunk não encontrado")

    elif menu == "Configurações":
        st.header("⚙️ Configurações")
        
        # Seção de configuração da API Key
        st.subheader("Configuração da API OpenAI")
        
        # Explica como configurar
        st.info("""
        Para usar recursos de IA, você precisa de uma chave API da OpenAI. 
        Você pode obtê-la em: https://platform.openai.com/api-keys
        """)
        
        # Campo para inserir a chave
        api_key = st.text_input("Chave da API OpenAI:", 
                              value=st.session_state.get('openai_api_key', ''), 
                              type="password",
                              help="Insira sua chave API da OpenAI aqui")
        
        if st.button("Salvar Chave API"):
            if api_key:
                st.session_state.openai_api_key = api_key
                st.success("✅ Chave API salva com sucesso!")
            else:
                st.warning("Por favor, insira uma chave API válida")
                
        # Seção de outras configurações
        st.subheader("Outras Configurações")
        new_chunk_size = st.number_input("Tamanho do chunk (caracteres):", 
                                       min_value=100, max_value=2000, 
                                       value=st.session_state.CHUNK_SIZE)
        
        if st.button("Atualizar Tamanho do Chunk"):
            st.session_state.CHUNK_SIZE = new_chunk_size
            st.success(f"✅ Tamanho do chunk atualizado para {st.session_state.CHUNK_SIZE} caracteres")

if __name__ == "__main__":
    main()